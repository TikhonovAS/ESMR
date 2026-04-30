import pandas as pd
from datetime import datetime, timedelta
import holidays
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os
from docxtpl import DocxTemplate

# 1. КОНСТАНТЫ И КАТАЛОГ РАБОТ
MONTH_NAMES_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}

WORK_CATALOG = {
    'vessel': {
        'label': 'Емкость/Резервуар', 'prio': 1,
        'ТР': 'Вскрытие, дегазация, очистка, дефектовка швов, толщинометрия, замена прокладок люков, испытания.',
        'ТО': 'Наружный осмотр, проверка дыхательной арматуры, контроль заземления.',
        'hours': {'ТР': 72, 'ТО': 8}, 'days': 3
    },
    'heatex': {
        'label': 'Теплообменник', 'prio': 2,
        'ТР': 'Разборка, механическая очистка трубных пучков, проверка на герметичность, замена уплотнений.',
        'ТО': 'Осмотр, контроль температуры/давления, проверка отсутствия протечек.',
        'hours': {'ТР': 64, 'ТО': 12}, 'days': 3
    },
    'pipe': {
        'label': 'Трубопровод', 'prio': 3,
        'ТР': 'Ревизия опорожнения, замена дефектных участков, проверка швов, восстановление изоляции.',
        'ТО': 'Обход трассы, проверка состояния опор, контроль герметичности.',
        'hours': {'ТР': 14, 'ТО': 2}, 'days': 1
    },
    'pump': {
        'label': 'Насос', 'prio': 4,
        'ТР': 'Центровка валов, разборка, замена подшипников и уплотнений, ревизия муфты.',
        'ТО': 'Проверка масла, контроль вибрации и шума, подтяжка болтов.',
        'hours': {'ТР': 18, 'ТО': 4}, 'days': 1
    },
    'fan': {
        'label': 'Вентилятор', 'prio': 5,
        'ТР': 'Балансировка лопастей, замена подшипников, проверка натяжения ремней.',
        'ТО': 'Очистка лопастей, контроль вибрации, проверка ограждений.',
        'hours': {'ТР': 16, 'ТО': 4}, 'days': 1
    },
    'loading': {
        'label': 'Устройство слива/налива', 'prio': 6,
        'ТР': 'Ревизия шарнирных соединений, замена уплотнений, проверка заземления.',
        'ТО': 'Проверка герметичности, смазка подвижных частей.',
        'hours': {'ТР': 14, 'ТО': 2}, 'days': 1
    },
    'bridge': {
        'label': 'Мостик переходной', 'prio': 7,
        'ТР': 'Ремонт ступеней, поручней, проверка узлов крепления, окраска.',
        'ТО': 'Осмотр конструкции, проверка надежности ограждений.',
        'hours': {'ТР': 12, 'ТО': 2}, 'days': 1
    }
}


# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
def get_tatarstan_holidays(year):
    ru_hols = holidays.Russia(years=[year, year + 1])
    rt_dates = [f"{year}-04-10", f"{year}-06-16", f"{year}-08-30", f"{year}-11-06",
                f"{year + 1}-03-30", f"{year + 1}-06-06"]
    all_hols = {d.strftime('%Y-%m-%d') for d in ru_hols}
    all_hols.update(rt_dates)
    return all_hols


HOLIDAYS_SET = set()


def is_workday(date):
    return date.weekday() < 5 and date.strftime('%Y-%m-%d') not in HOLIDAYS_SET


def get_end_date(start_date, duration_days):
    if duration_days <= 1: return start_date
    curr_date, counted = start_date, 1
    while counted < duration_days:
        curr_date += timedelta(days=1)
        if is_workday(curr_date): counted += 1
    return curr_date


def get_eq_cat(name):
    n = str(name).lower()
    if any(x in n for x in ['емкость', 'сосуд', 'резервуар']): return 'vessel'
    if 'теплообменник' in n: return 'heatex'
    if 'трубопровод' in n: return 'pipe'
    if 'насос' in n: return 'pump'
    if 'вентилятор' in n: return 'fan'
    if any(x in n for x in ['налив', 'слив']): return 'loading'
    if 'мостик' in n: return 'bridge'
    return 'pipe'


def get_all_workdays_in_month(year, month):
    workdays = []
    for d in range(1, 32):
        try:
            curr = datetime(year, month, d)
            if is_workday(curr): workdays.append(curr)
        except:
            break
    return workdays


# 2. РАСЧЕТ ГРАФИКА
def calculate_maintenance_balanced(df, target_year=2023):
    global HOLIDAYS_SET
    HOLIDAYS_SET = get_tatarstan_holidays(target_year)
    df.columns = [str(c).lower().strip() for c in df.columns]
    slots = defaultdict(int)

    def get_v(row, keys):
        for k in keys:
            for c in df.columns:
                if k in c:
                    val = row.get(c)
                    if pd.isna(val) or str(val).lower() == 'nan' or str(val).strip() == "": return "—"
                    return str(val).strip()
        return "—"

    data = df.to_dict('records')
    sorted_data = sorted(data, key=lambda x: (WORK_CATALOG[get_eq_cat(get_v(x, ['наименов']))]['prio'],
                                              get_v(x, ['наименов'])))

    res = []
    for i, row in enumerate(sorted_data):
        name = get_v(row, ['наименов'])
        cat_key = get_eq_cat(name)
        cfg = WORK_CATALOG[cat_key]

        m_tr = (i % 6) + 4
        w_days = get_all_workdays_in_month(target_year, m_tr)
        d_idx = slots[m_tr] % len(w_days)
        slots[m_tr] += 1
        start_tr = w_days[d_idx]

        jobs = {}
        jobs[m_tr] = {
            'date': start_tr, 'end_date': get_end_date(start_tr, cfg['days']),
            'type': 'ТР', 'hours': cfg['hours']['ТР'], 'desc': cfg['ТР'], 'day': start_tr.day
        }

        step = 3 if cat_key in ['pump', 'fan'] else 6
        for off in [-9, -6, -3, 3, 6, 9]:
            if abs(off) % step != 0: continue
            ideal = pd.Timestamp(start_tr) + pd.DateOffset(months=off)
            if (ideal.year == target_year and ideal.month >= 1) or (ideal.year == target_year + 1 and ideal.month <= 3):
                m_to = ideal.month
                if m_to not in jobs:
                    to_days = get_all_workdays_in_month(ideal.year, m_to)
                    to_start = to_days[d_idx % len(to_days)]
                    lbl = "ТО" if step == 6 else f"ТО-{ {3: 1, 6: 2, 9: 3, 12: 3}.get(off if off > 0 else 12 + off, 1)}"
                    jobs[m_to] = {
                        'date': to_start, 'end_date': to_start,
                        'type': lbl, 'hours': cfg['hours']['ТО'], 'desc': cfg['ТО'], 'day': to_start.day
                    }

        res.append({
            'name': name, 'marka': get_v(row, ['марка']),
            'zav_no': get_v(row, ['зав']), 'poz_no': get_v(row, ['поз']),
            'jobs': jobs, 'cat': cat_key
        })
    return res


# 3. ЭКСПОРТ EXCEL
def export_to_excel(schedule, output_path):
    if not os.path.exists('output'): os.makedirs('output', exist_ok=True)
    m_raw = [i if i <= 12 else i - 12 for i in range(4, 16)]
    m_names = [MONTH_NAMES_RU[m] for m in m_raw]
    brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        yr_rows = []
        for it in schedule:
            r = {'Оборудование': it['name'], 'Марка': it['marka'], 'Зав. №': it['zav_no'], 'Поз. №': it['poz_no']}
            for m in m_names: r[m] = ""
            for mi, j in it['jobs'].items():
                if MONTH_NAMES_RU[mi] in r: r[MONTH_NAMES_RU[mi]] = f"{j['type']} ({j['date'].strftime('%d.%m')})"
            yr_rows.append(r)
        pd.DataFrame(yr_rows).to_excel(writer, index=False, sheet_name='Годовой график')

        for mi in m_raw:
            m_n = MONTH_NAMES_RU[mi]
            y_v = 2024 if mi >= 4 else 2025  # Условный год для закраски выходных

            p_data, total_h = [], 0
            for it in schedule:
                if mi in it['jobs']:
                    j = it['jobs'][mi]
                    total_h += j['hours']
                    p_data.append({
                        'Начало': j['date'], 'Окончание': j['end_date'],
                        'Оборудование': it['name'], 'Марка': it['marka'],
                        'Зав. №': it['zav_no'], 'Поз. №': it['poz_no'],
                        'Вид работ': j['type'], 'Содержание работ': j['desc'], 'ч/час': j['hours']
                    })

            if p_data:
                df_p = pd.DataFrame(p_data).sort_values(['Начало', 'Вид работ'], ascending=[True, False])
                df_p['Начало'] = df_p['Начало'].dt.strftime('%d.%m.%Y')
                df_p['Окончание'] = df_p['Окончание'].dt.strftime('%d.%m.%Y')
                cols = ['Начало', 'Окончание', 'Оборудование', 'Марка', 'Зав. №', 'Поз. №', 'Вид работ',
                        'Содержание работ', 'ч/час']
                df_p = df_p[cols]

                sheet_n = f"План_{m_n}"
                df_p.to_excel(writer, index=False, sheet_name=sheet_n)
                writer.sheets[sheet_n].append(["", "", "", "", "", "", "", "ИТОГО ч/час:", total_h])

                ws_c = writer.book.create_sheet(title=f"Календарь_{m_n}")
                last_d = (pd.Timestamp(y_v, mi, 1) + pd.offsets.MonthEnd(0)).day
                ws_c.append(['Оборудование', 'Марка', 'Зав. №', 'Поз. №'] + list(range(1, last_d + 1)))
                for it in schedule:
                    if mi in it['jobs']:
                        j = it['jobs'][mi]
                        row = [it['name'], it['marka'], it['zav_no'], it['poz_no']] + [""] * last_d
                        row[j['day'] + 3] = j['type']
                        ws_c.append(row)
                for d in range(1, last_d + 1):
                    if not is_workday(datetime(y_v, mi, d)):
                        col_l = get_column_letter(d + 4)
                        for r_idx in range(1, ws_c.max_row + 1):
                            ws_c[f"{col_l}{r_idx}"].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                                                                       fill_type="solid")

        # --- НОВЫЙ БЛОК ОФОРМЛЕНИЯ ---
        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for r_idx, row in enumerate(ws.iter_rows()):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    cell.border = brd
                    if r_idx == 0 or "ИТОГО" in str(cell.value): cell.font = Font(bold=True)
                    if str(cell.value) == 'ТР':
                        cell.font = Font(bold=True, color="FF0000")
                    elif str(cell.value).startswith('ТО'):
                        cell.font = Font(bold=True, color="0000FF")

            # Используем перебор по индексам вместо объекта столбца
            for c_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(c_idx)
                if "Календарь" in sn and c_idx > 4:
                    ws.column_dimensions[col_letter].width = 4
                else:
                    max_l = 0
                    for row in ws.iter_rows(min_col=c_idx, max_col=c_idx):
                        for cell in row:
                            if cell.value:
                                curr_l = max(len(ln) for ln in str(cell.value).split('\n'))
                                if curr_l > max_l: max_l = curr_l
                    ws.column_dimensions[col_letter].width = min(max_l + 5, 50)

    print(f"График ППР успешно сформирован в папку output/")


# 4. ГЕНЕРАЦИЯ WORD
def generate_word_permits(schedule):
    template_path = 'templates/template_hazard.docx'
    output_dir = 'output/Наряды_Допуски'
    if not os.path.exists(output_dir): os.makedirs(output_dir, exist_ok=True)
    if not os.path.exists(template_path): return

    count = 0
    for it in schedule:
        if it['cat'] == 'vessel':
            for mi, j in it['jobs'].items():
                if j['type'] == 'ТР':
                    doc = DocxTemplate(template_path)
                    doc.render({
                        'date': j['date'].strftime('%d.%m.%Y'),
                        'end_date': j['end_date'].strftime('%d.%m.%Y'),
                        'name': it['name'], 'marka': it['marka'],
                        'zav_no': it['zav_no'], 'poz_no': it['poz_no'],
                        'hours': j['hours'], 'jobs_list': j['desc']
                    })
                    name_clean = str(it['name']).replace('/', '_').replace('\\', '_')
                    doc.save(os.path.join(output_dir, f"Наряд_{name_clean}_{j['date'].strftime('%d_%m_%Y')}.docx"))
                    count += 1
    if count > 0: print(f">>> Создано Word-нарядов: {count}")
